"""
Material Data Aggregation Script
=================================

Version: 1.0.0 
"""

import pandas as pd
import logging
from pathlib import Path
from typing import Dict, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import sys
from datetime import datetime
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ============================================================================
# Configuration
# ============================================================================

class Config:
    """Configuration settings for the aggregation process."""
    
    INPUT_FOLDER = "data"
    OUTPUT_FOLDER = "output"
    LOG_FOLDER = "logs"
    OUTPUT_FILENAME = "result.xlsx"
    
    INPUT_FILES = {
        'materials': 'materials.xlsx',
        'plants': 'plants.xlsx',
        'storage': 'storage.xlsx',
        'suppliers': 'suppliers.xlsx',
        'supplier_names': 'supplier-names.xlsx',
        'manufacturer_names': 'manufacturer-names.xlsx'
    }
    
    OUTPUT_COLUMNS = [
        'MaterialReference',
        'ManufacturerName',
        'ArticleNumber',
        'TypeCode',
        'ShortText',
        'Plant',
        'Disposition',
        'ReporderPoint',
        'SupplierName',
        'SupplierArticleNumber',
        'StorageLocation',
        'StorageBin',
        'DeletedStorageLevel'
    ]

# ============================================================================
# Logging Setup
# ============================================================================

def setup_logging():
    """Configure logging to file and console."""
    log_folder = Path(Config.LOG_FOLDER)
    log_folder.mkdir(parents=True, exist_ok=True)
    
    log_file = log_folder / f"aggregation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

logger = setup_logging()

# ============================================================================
# Data Loader
# ============================================================================

class DataLoader:
    """Handles loading Excel files from the input folder."""
    
    def __init__(self, input_folder: str):
        self.input_folder = Path(input_folder)
        
    def load_file(self, filename: str) -> Optional[pd.DataFrame]:
        """Load a single Excel file."""
        file_path = self.input_folder / filename
        
        try:
            if not file_path.exists():
                logger.error(f"File not found: {file_path}")
                return None
            
            df = pd.read_excel(file_path, engine='openpyxl')
            
            for col in df.columns:

                # Trim whitespace
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].replace('nan', pd.NA)

                # Preserve leading zeros for Plant
                if col == 'Plant':
                    df[col] = (
                        df[col]
                        .where(df[col].notna(), pd.NA)
                        .astype(str)
                        .str.replace(r'\.0$', '', regex=True)
                        .str.zfill(4)
                        .replace('000nan', pd.NA)
                    )

                if col in ['SupplierID', 'ManufacturerID']:
                    df[col] = (
                        df[col]
                        .where(df[col].notna(), pd.NA)
                        .astype(str)
                        .str.replace(r'\.0$', '', regex=True)
                        .str.strip()
                    )

            
            logger.info(f"Loaded {filename}: {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except Exception as e:
            logger.error(f"[ERROR] Error loading {filename}: {str(e)}")
            return None
    
    def load_all(self, file_mapping: Dict[str, str]) -> Tuple[Dict[str, pd.DataFrame], bool]:
        """Load all required files."""
        logger.info("=" * 70)
        logger.info("LOADING INPUT FILES")
        logger.info("=" * 70)
        
        dataframes = {}
        all_loaded = True
        # File Error handling
        for data_type, filename in file_mapping.items():
            df = self.load_file(filename)
            if df is None:
                all_loaded = False
                logger.error(f"Failed to load required file: {filename}")
            else:
                dataframes[data_type] = df
        
        if all_loaded:
            logger.info("All files loaded successfully\n")
        else:
            logger.error("[ERROR] Some files failed to load\n")
        
        return dataframes, all_loaded

# ============================================================================
# Data Aggregator
# ============================================================================

class MaterialDataAggregator:
    """Aggregates material data from multiple sources."""
    
    def __init__(self, dataframes: Dict[str, pd.DataFrame]):
        self.data = dataframes
        
    def get_primary_suppliers(self) -> pd.DataFrame:

        if 'suppliers' not in self.data:
            logger.warning("Suppliers data not available")
            return pd.DataFrame()
        
        suppliers = self.data['suppliers'].copy()
        
        suppliers['SupplierID_int'] = suppliers['SupplierID'].astype(int)
        
        # Sort by SupplierID and take first (lowest) per material
        suppliers_sorted = suppliers.sort_values('SupplierID_int')
        primary_suppliers = suppliers_sorted.groupby('MaterialReference').first().reset_index()
        
        primary_suppliers = primary_suppliers.drop('SupplierID_int', axis=1)
        
        logger.info(f"  Selected primary supplier (lowest ID) from {len(suppliers)} records → {len(primary_suppliers)} materials")
        return primary_suppliers
    
    def aggregate(self) -> pd.DataFrame:
        """Perform the complete aggregation process."""
        logger.info("=" * 70)
        logger.info("AGGREGATING DATA")
        logger.info("=" * 70)
        
        try:
            if 'storage' not in self.data:
                raise ValueError("Storage data is required but not found")
            
            result = self.data['storage'].copy()
            # logger.info(f"Step 1: Base table (storage) - {len(result)} rows")
            
            # Add materials data
            if 'materials' in self.data:
                before = len(result)
                result = result.merge(
                    self.data['materials'],
                    on='MaterialReference',
                    how='left'
                )
                # logger.info(f"Step 2: Merge materials - {before} → {len(result)} rows")
            
            # Add manufacturer names
            if 'manufacturer_names' in self.data:
                before = len(result)
                result = result.merge(
                    self.data['manufacturer_names'],
                    on='ManufacturerID',
                    how='left'
                )
                # logger.info(f"Step 3: Merge manufacturer names - {before} → {len(result)} rows")
            
            # Add plant data
            if 'plants' in self.data:
                before = len(result)
                result = result.merge(
                    self.data['plants'],
                    on=['MaterialReference', 'Plant'],
                    how='left'
                )
                # logger.info(f"Step 4: Merge plants - {before} → {len(result)} rows")
            
            # Add primary supplier data
            primary_suppliers = self.get_primary_suppliers()
            if not primary_suppliers.empty:
                before = len(result)
                result = result.merge(
                    primary_suppliers,
                    on='MaterialReference',
                    how='left'
                )
                # logger.info(f"Step 5: Merge primary suppliers - {before} → {len(result)} rows")
            
            # Add supplier names lookup
            if 'supplier_names' in self.data and 'SupplierID' in result.columns:
                before = len(result)
                result = result.merge(
                    self.data['supplier_names'],
                    on='SupplierID',
                    how='left'
                )
                # logger.info(f"Step 6: Add supplier names - {before} → {len(result)} rows")
            
            # Select and order columns
            for col in Config.OUTPUT_COLUMNS:
                if col not in result.columns:
                    result[col] = pd.NA
            
            result = result[Config.OUTPUT_COLUMNS]
            
            logger.info(f"\nAggregation complete: {len(result)} rows, {len(result.columns)} columns")
            logger.info("=" * 70)
            logger.info("")
            
            return result
            
        except Exception as e:
            logger.error(f"[ERROR] Aggregation failed: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            raise


# ============================================================================
# Data Validator
# ============================================================================

class DataValidator:
    """
    Validates both source data integrity and final output.
    """

    # Final key
    FINAL_GRAIN = [
        'MaterialReference',
        'Plant',
        'StorageLocation',
        'StorageBin'
    ]

    # Expected keys of source tables
    SOURCE_GRAINS = {
        'materials': ['MaterialReference'],
        'plants': ['MaterialReference', 'Plant'],
        'storage': ['MaterialReference', 'Plant', 'StorageLocation', 'StorageBin'],
        'suppliers': ['MaterialReference', 'SupplierID']
    }

    @staticmethod
    def validate_sources(dataframes: Dict[str, pd.DataFrame]) -> None:
        """
        Detect duplicate rows on expected source keys.
        """
        logger.info("=" * 70)
        logger.info("VALIDATING SOURCE DATA")
        logger.info("=" * 70)

        for name, keys in DataValidator.SOURCE_GRAINS.items():
            df = dataframes.get(name)
            if df is None:
                continue

            duplicates = df.duplicated(subset=keys, keep=False)
            if duplicates.any():
                logger.warning(
                    f"{name}: {duplicates.sum()} duplicate rows detected on key {keys}"
                )
            else:
                logger.info(f"{name}: no duplicates on key {keys}")

        logger.info("=" * 70)
        logger.info("")

    @staticmethod
    def validate_final(df: pd.DataFrame) -> bool:

        logger.info("=" * 70)
        logger.info("VALIDATING FINAL OUTPUT")
        logger.info("=" * 70)

        issues = []

        # Empty result check
        if df.empty:
            issues.append("Final result is empty")

        # Missing required columns
        missing_cols = [c for c in Config.OUTPUT_COLUMNS if c not in df.columns]
        if missing_cols:
            issues.append(f"Missing output columns: {missing_cols}")

        # NULLs in final grain
        null_keys = df[DataValidator.FINAL_GRAIN].isna().any(axis=1)
        if null_keys.any():
            issues.append(
                f"{null_keys.sum()} rows have NULLs in final grain keys "
                f"{DataValidator.FINAL_GRAIN}"
            )

        # Duplicate final grain rows
        duplicates = df.duplicated(subset=DataValidator.FINAL_GRAIN, keep=False)
        if duplicates.any():
            issues.append(
                f"{duplicates.sum()} Duplicate rows detected "
                f"{DataValidator.FINAL_GRAIN}"
            )

        if issues:
            logger.error("VALIDATION FAILED:")
            for issue in issues:
                logger.error(f"  - {issue}")
            return False

        logger.info("Validation PASSED")
        logger.info("=" * 70)
        logger.info("")
        return True



# ============================================================================
# Output Writer to Excel
# ============================================================================

class OutputWriter:
    """Handles writing the result to Excel file."""
    
    def __init__(self, output_folder: str):
        self.output_folder = Path(output_folder)
        
    def write(self, df: pd.DataFrame, filename: str) -> bool:
        """Write DataFrame to Excel file."""
        try:
            self.output_folder.mkdir(parents=True, exist_ok=True)
            output_path = self.output_folder / filename
            
            logger.info("=" * 70)
            logger.info("WRITING OUTPUT")
            logger.info("=" * 70)
            logger.info(f"Output file: {output_path}")
            
            df.to_excel(
                output_path,
                sheet_name='Aggregated Data',
                index=False,
                engine='openpyxl'
            )
            wb = load_workbook(output_path)
            ws = wb['Aggregated Data']

            plain_font = Font(bold=False)
            plain_alignment = Alignment(horizontal='left', vertical='bottom')

            for cell in ws[1]:
                cell.font = plain_font
                cell.alignment = plain_alignment

            wb.save(output_path)

            logger.info(f"Successfully wrote {len(df)} rows to {output_path}")
            logger.info("=" * 70)
            logger.info("")
            
            return True
            
        except Exception as e:
            logger.error(f"[ERROR] Failed to write output: {str(e)}")
            return False

# ============================================================================
# Main Pipeline
# ============================================================================

def main():
    """Execute the complete aggregation pipeline."""
    
    logger.info("")
    logger.info("=" * 70)
    logger.info("MATERIAL DATA AGGREGATION")
    logger.info("=" * 70)
    logger.info(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("")
    
    try:
        # Load all input files
        loader = DataLoader(Config.INPUT_FOLDER)
        dataframes, success = loader.load_all(Config.INPUT_FILES)
        
        if not success:
            logger.error("[ERROR] Failed to load all required files")
            return False
        
        DataValidator.validate_sources(dataframes)
        # Aggregate the data
        aggregator = MaterialDataAggregator(dataframes)
        result_df = aggregator.aggregate()
        

        if not DataValidator.validate_final(result_df):
            logger.ERROR("Validation found issues in final output, aborting...")
            return False

        # Write the output
        writer = OutputWriter(Config.OUTPUT_FOLDER)
        if not writer.write(result_df, Config.OUTPUT_FILENAME):
            logger.error("[ERROR] Failed to write output file")
            return False
        
        logger.info("")
        logger.info("=" * 70)
        logger.info("[SUCCESS] Aggregation Completed Successfully")
        logger.info("=" * 70)
        logger.info(f"End time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("")
        
        return True
        
    except Exception as e:
        logger.error("")
        logger.error("=" * 70)
        logger.error("[FAILED] AGGREGATION FAILED")
        logger.error("=" * 70)
        logger.error(f"Error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        logger.error("")
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)